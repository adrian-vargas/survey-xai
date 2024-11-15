import pandas as pd
import json
from pymongo import MongoClient
import os
from dotenv import load_dotenv
import matplotlib.pyplot as plt
import shutil
import numpy as np

# Cargar las variables de entorno desde el archivo .env
load_dotenv()

# Obtener las variables desde el entorno
mongo_uri = os.getenv('MONGO_URI')
db_name = os.getenv('MONGO_DB_NAME')
collection_name = os.getenv('MONGO_COLLECTION_NAME')

# Configuración de la conexión a MongoDB Atlas
client = MongoClient(mongo_uri)
db = client[db_name]
collection = db[collection_name]

# Cargar el archivo JSON 
with open('static/questions.json', 'r', encoding='utf-8') as f:
    questions_data = json.load(f).get("questions", [])

# Mapeo de valores para categorías, subcategorías, clases, preguntas de seguimiento y respuestas de seguimiento
category_map = {
    "Exactitud": 1,
    "Ambigüedad": 2,
    "Error": 3,
    "Preferencias de Visualización": 4,
    "Pregunta Descriptiva": 5
}

sub_category_map = {
    "Reglas": 1,
    "Grado Global": 2,
    "Grado Local": 3
}

class_map = {
    "Aprobado": 1,
    "Reprobado": 0
}

follow_up_question_map = {
    None: 0,
    "¿Qué tan seguro(a) estás de tu respuesta?": 1
}

follow_up_answer_map = {
    None: 0,
    "Nada": 1,
    "Poco": 2,
    "Mucho": 3
}

# Mapeo para normalizar los valores de user_answer
user_answer_map = {
    "Reprobado": 0,
    "Aprobado": 1,
    "Correcto": 2,
    "Incorrecto": 3,
    "Árbol de decisión (InterpretML)": 4,
    "Conjuntos de Decisiones interpretables (IDS)": 5,
    "Respuesta en texto libre": 6 # Para la respuesta de la pregunta 21 en formato libre
}

# Obtener todos los user_id únicos y asignar un número secuencial a cada uno
user_ids = collection.distinct("user_id")
user_id_map = {user_id: idx + 1 for idx, user_id in enumerate(user_ids)}

# Crear la estructura de datos para el glosario
glossary_data = []

# Añadir los valores de user_id al glosario
for user_id, normalized_id in user_id_map.items():
    glossary_data.append({"column": "user_id", "normalized_value": normalized_id, "original_value": user_id})

# Asignar un número secuencial a cada pregunta en el cuestionario y añadir al glosario
for question in questions_data:
    glossary_data.append({"column": "question", "normalized_value": question["id"], "original_value": question["instructions"]})

# Añadir los valores de category al glosario
for category_text, normalized_value in category_map.items():
    glossary_data.append({"column": "category", "normalized_value": normalized_value, "original_value": category_text})

# Añadir los valores de sub_category al glosario
for sub_category_text, normalized_value in sub_category_map.items():
    glossary_data.append({"column": "sub_category", "normalized_value": normalized_value, "original_value": sub_category_text})

# Añadir los valores de clases al glosario
for class_text, normalized_value in class_map.items():
    glossary_data.append({"column": "class", "normalized_value": normalized_value, "original_value": class_text})

# Añadir los valores de follow_up_question al glosario
for question_text, normalized_value in follow_up_question_map.items():
    glossary_data.append({"column": "follow_up_question", "normalized_value": normalized_value, "original_value": question_text})

# Añadir los valores de follow_up_answer al glosario
for answer_text, normalized_value in follow_up_answer_map.items():
    glossary_data.append({"column": "follow_up_answer", "normalized_value": normalized_value, "original_value": answer_text})

# Añadir los valores de user_answer al glosario
for answer_text, normalized_value in user_answer_map.items():
    glossary_data.append({"column": "user_answer", "normalized_value": normalized_value, "original_value": answer_text})

# Crear la carpeta 'report' si no existe
os.makedirs("report", exist_ok=True)

# Crear un archivo Excel con múltiples hojas
with pd.ExcelWriter("report/all_users_survey_report.xlsx") as writer:
    # Añadir la hoja del glosario con el formato solicitado
    glossary_df = pd.DataFrame(glossary_data)
    glossary_df.to_excel(writer, sheet_name="Glossary", index=False)

    # DataFrame para almacenar los conteos de respuestas generales
    general_counts = pd.DataFrame(index=range(1, 21))  # Para 20 preguntas
    general_counts.index.name = "question"  # Nombrar la primera columna como 'question'
    
    # Inicializar las columnas de conteo
    general_counts["reprobado"] = 0
    general_counts["aprobado"] = 0
    general_counts["correcto"] = 0
    general_counts["incorrecto"] = 0
    general_counts["dt"] = 0
    general_counts["ids"] = 0
    # Inicializar columnas para respuestas de seguimiento
    general_counts["aprobado_mucho"] = 0
    general_counts["aprobado_poco"] = 0
    general_counts["aprobado_nada"] = 0
    general_counts["reprobado_mucho"] = 0
    general_counts["reprobado_poco"] = 0
    general_counts["reprobado_nada"] = 0
    general_counts["correcto_mucho"] = 0
    general_counts["correcto_poco"] = 0
    general_counts["correcto_nada"] = 0
    general_counts["incorrecto_mucho"] = 0
    general_counts["incorrecto_poco"] = 0
    general_counts["incorrecto_nada"] = 0

    # Iterar sobre cada user_id para crear su reporte en una hoja individual
    for user_id, user_num in user_id_map.items():
        rows = []
        
        # Filtrar las respuestas correspondientes al user_id actual
        user_responses = list(collection.find({"user_id": user_id}))

        # Recorrer cada pregunta en el JSON y asignar las respuestas del usuario
        for idx, question in enumerate(questions_data[:21]):  # Limitar a 21 preguntas
            # Extraer datos de la pregunta y normalizar usando los mapeos
            question_id = question.get("id")
            category = question.get("category")
            normalized_category = category_map.get(category)
            sub_category = question.get("sub_category")
            normalized_sub_category = sub_category_map.get(sub_category)
            
            # Extraer observaciones
            observation = question.get("observation", {})
            absences = observation.get("absences")
            goout = observation.get("goout")
            studytime = observation.get("studytime")
            reason_reputation = observation.get("reason_reputation")
            failures = observation.get("failures")
            Fedu = observation.get("Fedu")
            
            # Obtener predicciones de los modelos y normalizarlas
            prediction_model = question.get("prediction_model", {})
            prediction_model_ids = class_map.get(prediction_model.get("IDS"))
            prediction_model_dt = class_map.get(prediction_model.get("DT-InterpretML"))
            
            # Normalizar la clase real
            real_prediction = class_map.get(question.get("real_class"))

            # Asociar datos de respuesta del usuario para la pregunta correspondiente
            response = user_responses[idx] if idx < len(user_responses) else {}
            user_answer_text = response.get('answer')
            
            # Normalizar `user_answer` para preguntas específicas (19, 20, y 21)
            if question_id in [19, 20]:
                print(f"Pregunta {question_id}: Respuesta de usuario original: {user_answer_text}")
                user_answer = user_answer_map.get(user_answer_text, user_answer_text)
                print(f"Pregunta {question_id}: Respuesta mapeada: {user_answer}")
            elif question_id == 21:
                user_answer = 6
                # Agregar la respuesta de la pregunta 21 al glosario
                glossary_data.append({
                    "column": f"user_id_{user_num}",
                    "normalized_value": 6,
                    "original_value": user_answer_text or ""
                })
            else:
                user_answer = user_answer_map.get(user_answer_text, user_answer_text)
            
            follow_up_answer = follow_up_answer_map.get(response.get('follow_up_answer'))
            
            # Convertir tiempo de milisegundos a segundos si está presente
            response_time_millis = response.get('response_time_seconds')
            response_time_seconds = response_time_millis / 1000 if response_time_millis else None
            
            # Pregunta de seguimiento normalizada
            follow_up_question = follow_up_question_map.get(question.get("follow_up", {}).get("question"))

            # Crear un diccionario con los datos de la fila con valores numéricos (normalizados) si no lo están
            row = {
                'user_id': user_num,  
                'question': question_id,  
                'category': normalized_category,  
                'sub_category': normalized_sub_category,
                'absences': absences,
                'goout': goout,
                'studytime': studytime,
                'reason_reputation': reason_reputation,
                'failures': failures,
                'Fedu': Fedu,
                'real_prediction': real_prediction,  
                'prediction_model_ids': prediction_model_ids,  
                'prediction_model_dt': prediction_model_dt,  
                'user_answer': user_answer,  
                'follow_up_question': follow_up_question,  
                'follow_up_answer': follow_up_answer,  
                'response_time_seconds': response_time_seconds  # Tiempo de respuesta en segundos
            }
            
            # Agregar la fila a la lista de filas del usuario actual
            rows.append(row)

            # Contar las respuestas en la tabla general
            # Contar las respuestas en la tabla general con impresión de depuración
            if user_answer == 0:
                general_counts.at[question_id, "reprobado"] += 1
                if follow_up_answer == 1:  # Nada
                    general_counts.at[question_id, "reprobado_nada"] += 1
                    print(f"Pregunta {question_id}: Incrementando 'reprobado_nada' para respuesta 'Nada'")
                elif follow_up_answer == 2:  # Poco
                    general_counts.at[question_id, "reprobado_poco"] += 1
                    print(f"Pregunta {question_id}: Incrementando 'reprobado_poco' para respuesta 'Poco'")
                elif follow_up_answer == 3:  # Mucho
                    general_counts.at[question_id, "reprobado_mucho"] += 1
                    print(f"Pregunta {question_id}: Incrementando 'reprobado_mucho' para respuesta 'Mucho'")

            elif user_answer == 1:
                general_counts.at[question_id, "aprobado"] += 1
                if follow_up_answer == 1:  # Nada
                    general_counts.at[question_id, "aprobado_nada"] += 1
                    print(f"Pregunta {question_id}: Incrementando 'aprobado_nada' para respuesta 'Nada'")
                elif follow_up_answer == 2:  # Poco
                    general_counts.at[question_id, "aprobado_poco"] += 1
                    print(f"Pregunta {question_id}: Incrementando 'aprobado_poco' para respuesta 'Poco'")
                elif follow_up_answer == 3:  # Mucho
                    general_counts.at[question_id, "aprobado_mucho"] += 1
                    print(f"Pregunta {question_id}: Incrementando 'aprobado_mucho' para respuesta 'Mucho'")

            elif user_answer == 2:
                general_counts.at[question_id, "correcto"] += 1
                if follow_up_answer == 1:  # Nada
                    general_counts.at[question_id, "correcto_nada"] += 1
                    print(f"Pregunta {question_id}: Incrementando 'correcto_nada' para respuesta 'Nada'")
                elif follow_up_answer == 2:  # Poco
                    general_counts.at[question_id, "correcto_poco"] += 1
                    print(f"Pregunta {question_id}: Incrementando 'correcto_poco' para respuesta 'Poco'")
                elif follow_up_answer == 3:  # Mucho
                    general_counts.at[question_id, "correcto_mucho"] += 1
                    print(f"Pregunta {question_id}: Incrementando 'correcto_mucho' para respuesta 'Mucho'")

            elif user_answer == 3:
                general_counts.at[question_id, "incorrecto"] += 1
                if follow_up_answer == 1:  # Nada
                    general_counts.at[question_id, "incorrecto_nada"] += 1
                    print(f"Pregunta {question_id}: Incrementando 'incorrecto_nada' para respuesta 'Nada'")
                elif follow_up_answer == 2:  # Poco
                    general_counts.at[question_id, "incorrecto_poco"] += 1
                    print(f"Pregunta {question_id}: Incrementando 'incorrecto_poco' para respuesta 'Poco'")
                elif follow_up_answer == 3:  # Mucho
                    general_counts.at[question_id, "incorrecto_mucho"] += 1
                    print(f"Pregunta {question_id}: Incrementando 'incorrecto_mucho' para respuesta 'Mucho'")

            elif user_answer == 4:
                general_counts.at[question_id, "dt"] += 1

            elif user_answer == 5:
                general_counts.at[question_id, "ids"] += 1

        # Crear un DataFrame con las filas del usuario actual
        df = pd.DataFrame(rows)

        # Escribir el DataFrame en una hoja del archivo Excel con el nombre "User_<número secuencial>"
        sheet_name = f"User_{user_num}"
        df.to_excel(writer, sheet_name=sheet_name, index=False)

    # Actualizar el glosario con las respuestas de la pregunta 21 de cada usuario
    glossary_df = pd.DataFrame(glossary_data)
    glossary_df.to_excel(writer, sheet_name="Glossary", index=False)

    # Guardar el conteo general en una nueva hoja
    general_counts.to_excel(writer, sheet_name="General", index=True)

print("Archivo 'report/all_users_survey_report.xlsx' creado exitosamente con cada usuario en una hoja separada, un glosario de mapeos y un reporte general.")

################################# GRÁFICAS POR PREGUNTA PRINCIPAL #######################################

# Leer el archivo JSON para obtener los datos de las preguntas
with open('static/questions.json', 'r', encoding='utf-8') as f:
    questions_data = json.load(f).get("questions", [])

# Crear un DataFrame a partir del JSON para tener la relación entre cada pregunta y su categoría/subcategoría
questions_df = pd.DataFrame(questions_data)

# Leer el archivo Excel
file_path = 'report/all_users_survey_report.xlsx'
general_counts = pd.read_excel(file_path, sheet_name='General', index_col=0)

# Validar que el DataFrame general_counts tenga las columnas de categoría, subcategoría y las instrucciones
general_counts = general_counts.merge(
    questions_df[['id', 'category', 'sub_category', 'model', 'instructions', 'prediction_model', 'real_class', 'observation']],
    left_index=True,
    right_on='id',
    how='left'
)

# Crear carpetas para las categorías
categories = general_counts['category'].unique()
sub_categories = general_counts['sub_category'].unique()

for category in categories:
    # Crear carpeta para cada categoría
    folder_path = f'report/{category}'
    os.makedirs(folder_path, exist_ok=True)

# Definir los tipos de respuestas por categoría
response_options = {
    "Exactitud": ['aprobado', 'reprobado'],
    "Ambigüedad": ['aprobado', 'reprobado'],
    "Error": ['correcto', 'incorrecto']
}

# Generar gráficas para cada subcategoría dentro de una categoría
for category in categories:
    for sub_category in sub_categories:
        # Filtrar las preguntas que pertenecen a la categoría y subcategoría actual
        filtered_data = general_counts[
            (general_counts['category'] == category) &
            (general_counts['sub_category'] == sub_category)
        ]

        if not filtered_data.empty:
            plt.figure(figsize=(12, 8))  # Aumenta el tamaño si es necesario

            # Obtener los IDs y las instrucciones de las preguntas para agregar contexto
            question_ids = filtered_data['id'].tolist()
            instructions = filtered_data['instructions'].tolist()
            question_labels = [f"ID {qid}: {instr}" for qid, instr in zip(question_ids, instructions)]

            # Obtener los tipos de respuesta correspondientes a la categoría
            response_types = response_options.get(category, [])
            x = np.arange(len(response_types))  # Posiciones para cada tipo de respuesta en el eje X
            width = 0.35  # Ancho de las barras

            # Obtener la observación (es la misma para las preguntas de la misma subcategoría)
            observation = filtered_data['observation'].iloc[0]

            # Preparar el texto de la observación para mostrar en la gráfica
            observation_text = ', '.join([f"{key}: {value}" for key, value in observation.items()])

            # Iterar sobre cada modelo (DT-InterpretML y IDS)
            models = filtered_data['model'].unique()

            for idx, model in enumerate(models):
                model_data = filtered_data[filtered_data['model'] == model]

                # Calcular el total de respuestas para cada tipo de respuesta
                counts = [model_data[response_type].sum() for response_type in response_types]

                # Graficar las barras para el modelo actual
                plt.bar(x + idx * width, counts, width, label=f"{model}")

            # Agregar título y contexto a la gráfica
            plt.title(f'Conteo Total de Respuestas para {category} - {sub_category}')
            plt.xlabel('Opciones de Respuesta', labelpad=5)  # Ajuste del espacio con labelpad
            plt.ylabel('Número de Usuarios', labelpad=5)  # Ajuste del espacio con labelpad
            plt.xticks(x + width / 2, response_types, rotation=0)  # Alineación horizontal de etiquetas
            plt.legend()

            # Preparar datos para la tabla
            table_data = []
            for _, row in filtered_data.iterrows():
                table_data.append([
                    f"ID {row['id']}",
                    row['model'],
                    row['prediction_model'][row['model']],
                    row['real_class']
                ])

            # Crear una tabla en la parte inferior de la gráfica
            column_labels = ["Pregunta ID", "Modelo", "Predicción del Modelo", "Clase Real"]
            plt.table(
                cellText=table_data,
                colLabels=column_labels,
                cellLoc='center',
                loc='bottom',
                bbox=[0.0, -0.4, 1.0, 0.2]  # Ajuste de la posición y tamaño de la tabla
            )

            # Agregar la instrucción y observación debajo de la tabla
            instruction = '\n'.join(set(instructions))

            plt.figtext(
                0.5,
                0.1,
                f"Instrucción: {instruction}",
                wrap=True,
                horizontalalignment='center',
                fontsize=10,
                weight='normal'
            )
            plt.figtext(
                0.5,
                0.08,
                f"Observación: {observation_text}",
                wrap=True,
                horizontalalignment='center',
                fontsize=9
            )
            plt.subplots_adjust(left=0.1, right=0.9, top=0.9, bottom=0.2)

            # Guardar la gráfica en la carpeta correspondiente
            plt.savefig(f'report/{category}/{category}_{sub_category}_Totales.png', bbox_inches='tight')
            plt.close()

################################# GRÁFICAS POR PREGUNTA DE SEGUIMIENTO #######################################

# Leer el archivo JSON para obtener los datos de las preguntas
with open('static/questions.json', 'r', encoding='utf-8') as f:
    questions_data = json.load(f).get("questions", [])

# Crear un DataFrame a partir del JSON para tener la relación entre cada pregunta y su categoría/subcategoría
questions_df = pd.DataFrame(questions_data)

# Leer el archivo Excel
file_path = 'report/all_users_survey_report.xlsx'
general_counts = pd.read_excel(file_path, sheet_name='General', index_col=0)

# Validar que el DataFrame general_counts tenga las columnas de categoría, subcategoría y las instrucciones
general_counts = general_counts.merge(
    questions_df[['id', 'category', 'sub_category', 'model', 'instructions', 'prediction_model', 'real_class', 'observation']],
    left_index=True,
    right_on='id',
    how='left'
)

# Definir los tipos de respuestas para preguntas de seguimiento por categoría
follow_up_response_options = {
    "Ambigüedad": [
        'aprobado_mucho', 'aprobado_poco', 'aprobado_nada',
        'reprobado_mucho', 'reprobado_poco', 'reprobado_nada'
    ],
    "Error": [
        'correcto_mucho', 'correcto_poco', 'correcto_nada',
        'incorrecto_mucho', 'incorrecto_poco', 'incorrecto_nada'
    ]
}

# Generar gráficas para las preguntas de seguimiento en cada categoría
for category in ["Ambigüedad", "Error"]:
    # Crear carpeta para la categoría actual
    folder_path = f'report/follow_up_question/{category}'
    os.makedirs(folder_path, exist_ok=True)

    for sub_category in general_counts['sub_category'].unique():
        # Filtrar las preguntas que pertenecen a la categoría y subcategoría actual
        filtered_data = general_counts[
            (general_counts['category'] == category) & (general_counts['sub_category'] == sub_category)
        ]

        if not filtered_data.empty:
            plt.figure(figsize=(12, 8))

            # Obtener los tipos de respuesta de seguimiento correspondientes a la categoría
            follow_up_types = follow_up_response_options.get(category, [])
            x = np.arange(len(follow_up_types))  # Posiciones para cada tipo de respuesta en el eje X
            width = 0.35  # Ancho de las barras

            # Obtener la observación y la instrucción (es la misma para las preguntas de la misma subcategoría)
            observation = filtered_data['observation'].iloc[0]
            instruction = filtered_data['instructions'].iloc[0]

            # Preparar el texto de la observación para mostrar en la gráfica
            observation_text = ', '.join([f"{key}: {value}" for key, value in observation.items()])

            # Iterar sobre cada modelo (DT-InterpretML y IDS)
            models = filtered_data['model'].unique()

            for idx, model in enumerate(models):
                model_data = filtered_data[filtered_data['model'] == model]

                # Calcular el total de respuestas para cada tipo de respuesta de seguimiento
                counts = [model_data[follow_up_type].sum() for follow_up_type in follow_up_types]

                # Graficar las barras para el modelo actual
                plt.bar(x + idx * width, counts, width, label=f"{model}")

            plt.title(f'Conteo Total de Respuestas de Seguimiento para {category} - {sub_category}')
            plt.xlabel('Opciones de Respuesta de Seguimiento', labelpad=5)  # Ajuste del espacio con labelpad
            plt.ylabel('Número de Usuarios', labelpad=5)  # Ajuste del espacio con labelpad
            plt.xticks(x + width / 2, follow_up_types, rotation=0)  # Alineación horizontal de etiquetas
            plt.legend()

            # Preparar datos para la tabla
            table_data = []
            for _, row in filtered_data.iterrows():
                table_data.append([
                    f"ID {row['id']}",
                    row['model'],
                    row['prediction_model'][row['model']],
                    row['real_class']
                ])

            # Crear una tabla en la parte inferior de la gráfica
            column_labels = ["Pregunta ID", "Modelo", "Predicción del Modelo", "Clase Real"]
            plt.table(
                cellText=table_data,
                colLabels=column_labels,
                cellLoc='center',
                loc='bottom',
                bbox=[0.0, -0.4, 1.0, 0.2]  # Ajuste de la posición y tamaño de la tabla
            )

            # Agregar la instrucción y observación debajo de la tabla como texto adicional con ajustes en las coordenadas
            plt.figtext(
                0.5,
                0.1,
                f"Instrucción: {instruction}",
                wrap=True,
                horizontalalignment='center',
                fontsize=10,
                weight='normal'
            )
            plt.figtext(
                0.5,
                0.08,
                f"Observación: {observation_text}",
                wrap=True,
                horizontalalignment='center',
                fontsize=9
            )
            plt.subplots_adjust(left=0.1, right=0.9, top=0.9, bottom=0.2)

            # Guardar la gráfica en la carpeta correspondiente
            plt.savefig(
                f'{folder_path}/{category}_{sub_category}_Seguimiento_Totales.png',
                bbox_inches='tight'
            )
            plt.close()

################################# DESCARGA DEL REPORTE #######################################

# Ruta de la carpeta report y la ubicación de destino del ZIP
report_path = "report"
zip_path = "report.zip"
static_zip_path = "static/report.zip"

# Crear un archivo ZIP de la carpeta 'report' en su ubicación actual
shutil.make_archive(zip_path.replace('.zip', ''), 'zip', report_path)

# Mover el archivo ZIP a la carpeta static
if not os.path.exists("static"):
    os.makedirs("static")
shutil.move(zip_path, static_zip_path)
print("Archivo ZIP creado y movido a 'static/report.zip'")


################################# FORMATO DE REPORTE ##########################################
import openpyxl
from openpyxl.styles import PatternFill

# Cargar el archivo Excel generado
file_path = 'report/all_users_survey_report.xlsx'
workbook = openpyxl.load_workbook(file_path)

# Obtener la hoja necesaria
sheet = workbook['General']  # Asumiendo que los datos están en la hoja "General"

# Colores de formato
gray_fill_exactitud = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')  # Gris para exactitud
gray_fill_ambiguidad = PatternFill(start_color='BFBFBF', end_color='BFBFBF', fill_type='solid')  # Gris para ambigüedad
gray_fill_error = PatternFill(start_color='A6A6A6', end_color='A6A6A6', fill_type='solid')  # Gris para error
gray_fill_19_20 = PatternFill(start_color='9E9E9E', end_color='9E9E9E', fill_type='solid')  # Gris para preguntas 19 y 20 con valor 0

green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')  # Verde para "aprobado", "correcto"
red_fill = PatternFill(start_color='F4CCCC', end_color='F4CCCC', fill_type='solid')  # Rojo para "reprobado", "incorrecto"
blue_fill = PatternFill(start_color='C9DAF8', end_color='C9DAF8', fill_type='solid')  # Azul para "dt"
violet_fill = PatternFill(start_color='EAD1DC', end_color='EAD1DC', fill_type='solid')  # Violeta para "ids"

# Definir el orden correcto de las columnas
correct_column_order = [
    'question', 'aprobado', 'reprobado', 'aprobado_mucho', 'aprobado_poco', 'aprobado_nada',
    'reprobado_mucho', 'reprobado_poco', 'reprobado_nada', 'correcto_mucho', 'correcto_poco', 'correcto_nada',
    'incorrecto_mucho', 'incorrecto_poco', 'incorrecto_nada', 'correcto', 'incorrecto', 'ids', 'dt'
]

# Obtener los encabezados actuales de la hoja
columns = [cell.value for cell in sheet[1]]  # Obtener los encabezados de la primera fila
column_index_map = {column: index for index, column in enumerate(columns)}

# Crear una lista de los índices correspondientes a las columnas en el orden correcto
new_column_order = [column_index_map[col] for col in correct_column_order if col in column_index_map]

# Crear una nueva hoja para el reporte con las columnas en el orden correcto
new_sheet = workbook.create_sheet("Reordered General")

# Copiar los encabezados en el nuevo orden
for idx, col in enumerate(new_column_order, start=1):
    new_sheet.cell(row=1, column=idx, value=columns[col])

# Copiar los valores de las filas en el nuevo orden
for row_idx, row in enumerate(sheet.iter_rows(min_row=2, min_col=1, max_col=sheet.max_column), start=2):
    for col_idx, new_col in enumerate(new_column_order, start=1):
        new_sheet.cell(row=row_idx, column=col_idx, value=row[new_col].value)

# Aplicar los colores correspondientes a la nueva hoja con el orden corregido
for row in new_sheet.iter_rows(min_row=2, min_col=1, max_col=new_sheet.max_column):
    question = row[0].value  # La columna 'question' está en la columna A

    # Aplicar formato a la celda de la columna 'question' dependiendo de la categoría
    if question <= 6:  # Exactitud (Preguntas 1 a 6)
        row[0].fill = gray_fill_exactitud
    elif question <= 12:  # Ambigüedad (Preguntas 7 a 12)
        row[0].fill = gray_fill_ambiguidad
    elif question <= 18:  # Error (Preguntas 13 a 18)
        row[0].fill = gray_fill_error
    elif question == 19 or question == 20:  # Preguntas 19 y 20
        row[0].fill = gray_fill_19_20  # Gris más oscuro

    # Aplicar formato a las otras celdas dependiendo de sus valores y la columna a la que pertenecen
    for cell_idx, cell in enumerate(row[1:], start=1):  # Empezamos en la columna 2 (ya que la primera es 'question')
        header = new_sheet.cell(row=1, column=cell_idx + 1).value

        if cell.value == 0:
            if question <= 6:  # Exactitud (Preguntas 1 a 6)
                cell.fill = gray_fill_exactitud
            elif question <= 12:  # Ambigüedad (Preguntas 7 a 12)
                cell.fill = gray_fill_ambiguidad
            elif question <= 18:  # Error (Preguntas 13 a 18)
                cell.fill = gray_fill_error
            elif question == 19 or question == 20:  # Preguntas 19 y 20
                cell.fill = gray_fill_19_20  # Gris más oscuro si el valor es 0

        elif cell.value > 0:
            # Aplicar color verde para "aprobado", "correcto"
            if header in ['aprobado', 'correcto', 'aprobado_mucho', 'aprobado_poco', 'aprobado_nada', 'correcto_mucho', 'correcto_poco', 'correcto_nada']:
                cell.fill = green_fill
            
            # Aplicar color rojo para "reprobado", "incorrecto"
            elif header in ['reprobado', 'reprobado_mucho', 'reprobado_poco', 'reprobado_nada', 'incorrecto', 'incorrecto_mucho', 'incorrecto_poco', 'incorrecto_nada']:
                cell.fill = red_fill
            
            # Aplicar color azul para "dt" en preguntas 19 y 20
            elif header == 'dt' and (question == 19 or question == 20):
                cell.fill = blue_fill

            # Aplicar color violeta para "ids" en preguntas 19 y 20
            elif header == 'ids' and (question == 19 or question == 20):
                cell.fill = violet_fill

# Guardar el archivo después de aplicar el formato y reordenar las columnas
workbook.save('report/formatted_and_reordered_users_survey_report.xlsx')

print("Formato aplicado, columnas reorganizadas y archivo guardado como 'formatted_and_reordered_users_survey_report.xlsx'")
